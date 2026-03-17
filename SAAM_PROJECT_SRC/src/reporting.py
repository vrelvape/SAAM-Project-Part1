import numpy as np
import pandas as pd


def build_performance_table(mv_returns_oos, vw_returns_oos):
    """
    Build the performance comparison table for the minimum variance portfolio
    and the value-weighted benchmark.
    """

    mv_ann_return = mv_returns_oos.mean() * 12
    vw_ann_return = vw_returns_oos.mean() * 12

    mv_ann_vol = mv_returns_oos.std() * np.sqrt(12)
    vw_ann_vol = vw_returns_oos.std() * np.sqrt(12)

    mv_sharpe = mv_ann_return / mv_ann_vol
    vw_sharpe = vw_ann_return / vw_ann_vol

    mv_cum_return = (1 + mv_returns_oos).prod() - 1
    vw_cum_return = (1 + vw_returns_oos).prod() - 1

    performance = pd.DataFrame({
        "Portfolio": [
            "Minimum Variance",
            "Value Weighted"
        ],
        "Annualized Return": [
            mv_ann_return,
            vw_ann_return
        ],
        "Annualized Volatility": [
            mv_ann_vol,
            vw_ann_vol
        ],
        "Sharpe Ratio": [
            mv_sharpe,
            vw_sharpe
        ],
        "Minimum Monthly Return": [
            mv_returns_oos.min(),
            vw_returns_oos.min()
        ],
        "Maximum Monthly Return": [
            mv_returns_oos.max(),
            vw_returns_oos.max()
        ],
        "Total Cumulative Return": [
            mv_cum_return,
            vw_cum_return
        ]
    })

    return performance

def compute_cumulative_series(mv_returns_oos, vw_returns_oos):
    """
    Compute cumulative wealth series for the minimum variance portfolio
    and the value-weighted benchmark.
    """

    mv_cumulative = (1 + mv_returns_oos).cumprod()
    vw_cumulative = (1 + vw_returns_oos).cumprod()

    return mv_cumulative, vw_cumulative

import matplotlib.pyplot as plt

def plot_cumulative_performance(mv_cumulative, vw_cumulative, figures_dir, show_plot=True):
    """
    Plot and save the cumulative performance figure.
    """

    plt.figure(figsize=(10, 6))

    plt.plot(mv_cumulative, label="Minimum Variance", linewidth=2)
    plt.plot(vw_cumulative, label="Value Weighted", linewidth=2)

    plt.title("Cumulative Portfolio Performance (2014–2025)")
    plt.ylabel("Growth of $1 Investment")
    plt.xlabel("Date")

    plt.legend()
    plt.grid(True)

    output_path = figures_dir / "cumulative_portfolio_performance.png"
    plt.savefig(output_path, dpi=300, bbox_inches="tight")

    if show_plot:
        plt.show()
    else:
        plt.close()

    return output_path

def export_part1_outputs(mv_returns_oos, vw_returns_oos, performance, tables_dir):
    """
    Export key Part I outputs to CSV files.
    """

    mv_path = tables_dir / "mv_returns_oos.csv"
    vw_path = tables_dir / "vw_returns_oos.csv"
    performance_path = tables_dir / "portfolio_performance_summary.csv"

    mv_returns_oos.to_csv(mv_path)
    vw_returns_oos.to_csv(vw_path)
    performance.to_csv(performance_path, index=False)

    return {
        "mv_returns_oos": mv_path,
        "vw_returns_oos": vw_path,
        "performance_summary": performance_path,
    }

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import pandas as pd


def annualized_cumulative_return(return_series):
    """
    Compute annualized cumulative return from a monthly return series.
    """
    n_months = len(return_series.dropna())
    total_cum_return = (1 + return_series).prod() - 1
    return (1 + total_cum_return) ** (12 / n_months) - 1


def save_single_cumulative_plot(series, title, output_path):
    """
    Save a single cumulative performance plot for Excel template insertion.
    """
    plt.figure(figsize=(5, 3.2))
    plt.plot(series, linewidth=2)
    plt.title(title)
    plt.ylabel("Growth of $1")
    plt.xlabel("Date")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close()


from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import pandas as pd


def annualized_cumulative_return(return_series):
    """
    Compute annualized cumulative return from a monthly return series.
    """
    n_months = len(return_series.dropna())
    total_cum_return = (1 + return_series).prod() - 1
    return (1 + total_cum_return) ** (12 / n_months) - 1


def save_single_cumulative_plot(series, title, output_path):
    """
    Save a single cumulative performance plot for Excel template insertion.
    """
    plt.figure(figsize=(5, 3.2))
    plt.plot(series, linewidth=2)
    plt.title(title)
    plt.ylabel("Growth of $1")
    plt.xlabel("Date")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close()


from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import pandas as pd


def annualized_cumulative_return(return_series):
    """
    Compute annualized cumulative return from a monthly return series.
    """
    n_months = len(return_series.dropna())
    total_cum_return = (1 + return_series).prod() - 1
    return (1 + total_cum_return) ** (12 / n_months) - 1


def save_single_cumulative_plot(series, title, output_path):
    """
    Save a single cumulative performance plot for Excel template insertion.
    """
    plt.figure(figsize=(5, 3.2))
    plt.plot(series, linewidth=2)
    plt.title(title)
    plt.ylabel("Growth of $1")
    plt.xlabel("Date")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig(output_path, dpi=200, bbox_inches="tight")
    plt.close()


def fill_part1_excel_template(
    templates_dir,
    excel_dir,
    figures_dir,
    mv_returns_oos,
    vw_returns_oos,
    template_filename="Template for Part I-SAAM.xlsx",
    output_filename="Part_I_template_filled.xlsx",
):
    """
    Fill the official Part I Excel template with portfolio statistics,
    monthly returns, and cumulative performance plots.
    """

    template_path = templates_dir / template_filename
    filled_template_path = excel_dir / output_filename

    vw_plot_path = figures_dir / "vw_cumulative_template_plot.png"
    mv_plot_path = figures_dir / "mv_cumulative_template_plot.png"

    # Build cumulative series
    mv_cumulative = (1 + mv_returns_oos).cumprod()
    vw_cumulative = (1 + vw_returns_oos).cumprod()

    # Save two individual plots
    save_single_cumulative_plot(
        vw_cumulative,
        "Value-Weighted Portfolio",
        vw_plot_path
    )

    save_single_cumulative_plot(
        mv_cumulative,
        "Minimum Variance Portfolio",
        mv_plot_path
    )

    # Load workbook
    wb = load_workbook(template_path)
    ws = wb["Sheet1"]

    # Summary statistics
    vw_ann_return = vw_returns_oos.mean() * 12
    mv_ann_return = mv_returns_oos.mean() * 12

    vw_ann_vol = vw_returns_oos.std() * (12 ** 0.5)
    mv_ann_vol = mv_returns_oos.std() * (12 ** 0.5)

    vw_ann_cum = annualized_cumulative_return(vw_returns_oos)
    mv_ann_cum = annualized_cumulative_return(mv_returns_oos)

    vw_sharpe = vw_ann_return / vw_ann_vol
    mv_sharpe = mv_ann_return / mv_ann_vol

    vw_min = vw_returns_oos.min()
    mv_min = mv_returns_oos.min()

    vw_max = vw_returns_oos.max()
    mv_max = mv_returns_oos.max()

    # Fill portfolio characteristics
    ws["B3"] = vw_ann_return
    ws["C3"] = mv_ann_return

    ws["B4"] = vw_ann_vol
    ws["C4"] = mv_ann_vol

    ws["B5"] = vw_ann_cum
    ws["C5"] = mv_ann_cum

    ws["B6"] = vw_sharpe
    ws["C6"] = mv_sharpe

    ws["B7"] = vw_min
    ws["C7"] = mv_min

    ws["B8"] = vw_max
    ws["C8"] = mv_max

    # Fill monthly returns
    vw_map = {pd.Timestamp(d).to_period("M"): v for d, v in vw_returns_oos.items()}
    mv_map = {pd.Timestamp(d).to_period("M"): v for d, v in mv_returns_oos.items()}

    for row in range(3, ws.max_row + 1):
        cell_date = ws[f"E{row}"].value

        if isinstance(cell_date, pd.Timestamp):
            period = cell_date.to_period("M")
        elif hasattr(cell_date, "year") and hasattr(cell_date, "month"):
            period = pd.Timestamp(cell_date).to_period("M")
        else:
            continue

        ws[f"F{row}"] = vw_map.get(period, None)
        ws[f"G{row}"] = mv_map.get(period, None)

    # Insert plots
    img_vw = XLImage(str(vw_plot_path))
    img_mv = XLImage(str(mv_plot_path))

    img_vw.width = 260
    img_vw.height = 170

    img_mv.width = 260
    img_mv.height = 170

    ws.add_image(img_vw, "B9")
    ws.add_image(img_mv, "C9")

    # Save final workbook
    wb.save(filled_template_path)

    return filled_template_path